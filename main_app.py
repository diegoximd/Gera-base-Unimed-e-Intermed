import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import requests
import json
import pandas as pd
from datetime import datetime, timedelta
import sys
import os
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from tkcalendar import DateEntry
import locale

# IMPORTAÇÃO DOS DADOS DE CONFIGURAÇÃO E CREDENCIAIS
# Todos os dicionários e listas (API_SECRETS, MAP_COLUMNS, etc.) são importados daqui.
import config

# Configura o locale
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252')
    except locale.Error:
        print("Aviso: Não foi possível configurar o locale PT-BR.")


# --- FUNÇÕES DE UTILIDADE ---

def load_credentials():
    """Retorna as credenciais de API diretamente do módulo config."""
    # Checagem básica para garantir que o desenvolvedor preencheu os campos
    if any(v.startswith("[URL_DE_AUTENTICACAO") for k, v in config.API_SECRETS["unimed"].items() if isinstance(v, str)):
        messagebox.showerror("Erro de Configuração",
                             "As credenciais no dicionário API_SECRETS no arquivo config.py não foram preenchidas com dados reais.")
        return None

    return config.API_SECRETS


def calculate_dates():
    """Calcula vencimentoFinal (120 dias antes da requisição) e vencimentoInicio (1 semana antes)."""
    hoje = datetime.now()
    venc_final_dt = hoje - timedelta(days=120)
    venc_inicio_dt = venc_final_dt - timedelta(weeks=1)
    return (
        venc_inicio_dt.strftime("%d/%m/%Y"),
        venc_final_dt.strftime("%d/%m/%Y")
    )


# --- Funções de API ---

def authenticate_api(empresa_key):
    """Autentica na API e retorna o token Bearer."""
    creds = load_credentials()
    if not creds: return None

    empresa_data = creds.get(empresa_key)
    if not empresa_data:
        messagebox.showerror("Erro", f"Dados da empresa '{empresa_key}' não encontrados no config.API_SECRETS.")
        return None

    auth_url = empresa_data.get('auth_url')
    body = {
        "usuario": empresa_data.get('usuario'),
        "senha": empresa_data.get('senha'),
        "client_id": empresa_data.get('client_id')
    }

    headers = {
        "Content-Type": "application/json",
        "Connection": "keep-alive",
        "Origin": "http://179.189.118.34"
    }

    try:
        response = requests.post(auth_url, json=body, headers=headers, timeout=10)
        response.raise_for_status()
        token = response.json().get('token')
        if not token:
            messagebox.showerror("Erro de Autenticação",
                                 "Resposta da API não contém o token. Verifique usuário/senha/client_id.")
        return token

    except requests.exceptions.HTTPError as e:
        messagebox.showerror("Erro de Autenticação", f"Falha HTTP: {e}\nCorpo da Resposta: {response.text[:200]}...")
        return None
    except requests.exceptions.RequestException as e:
        messagebox.showerror("Erro de Conexão", f"Falha ao conectar à API de autenticação: {e}")
        return None
    except json.JSONDecodeError:
        messagebox.showerror("Erro de Dados",
                             f"Resposta da API não é um JSON válido. Resposta: {response.text[:100]}...")
        return None


def fetch_api_data(empresa_key, vencimento_inicio, vencimento_final, baixa_pdd, token):
    """Acessa o endpoint de dados da API e retorna os dados brutos."""
    creds = load_credentials()
    if not creds: return None

    empresa_data = creds.get(empresa_key)
    data_url = empresa_data.get('data_url')

    if not data_url:
        messagebox.showerror("Erro de Configuração", "URL de dados da API ausente no config.API_SECRETS.")
        return None

    params = {
        "vencimentoInicio": vencimento_inicio,
        "vencimentoFinal": vencimento_final,
        "baixaPdd": baixa_pdd
    }
    headers = {
        "Content-Type": "application/json",
        "Connection": "keep-alive",
        "Authorization": f"Bearer {token}",
        "Origin": "http://179.189.118.34"
    }

    try:
        response = requests.get(data_url, params=params, headers=headers, timeout=30)
        response.raise_for_status()

        data = response.json()
        return data

    except requests.exceptions.HTTPError as e:
        messagebox.showerror("Erro de Dados",
                             f"Falha HTTP ao buscar dados: {e}\nCorpo da Resposta: {response.text[:200]}...")
        return None
    except requests.exceptions.RequestException as e:
        messagebox.showerror("Erro de Conexão", f"Falha ao buscar dados da API: {e}")
        return None
    except json.JSONDecodeError:
        messagebox.showerror("Erro de Dados",
                             f"Resposta da API não é um JSON válido. Resposta: {response.text[:100]}...")
        return None


# --- Processamento de Dados ---

def process_data_no_db(api_data):
    """
    Transforma dados da API, aplica formatação monetária e garante a estrutura de 113 colunas.
    """
    all_records = []
    operations_by_cpf = api_data.get('data', {})

    for _, operations in operations_by_cpf.items():
        if isinstance(operations, list):
            for op in operations:
                if isinstance(op, dict):

                    # 1. TRATAMENTO DO CAMPO OBS. OPERAÇÃO (Beneficiário)
                    benef_data = op.get('benefs_contrato', '')
                    if benef_data:
                        op['benefs_contrato'] = f"Beneficiário: {benef_data}"

                    # 2. FORMATAÇÃO MONETÁRIA
                    for api_key in ['vl_venda', 'vl_vencido']:
                        value = op.get(api_key)
                        if value is not None:
                            try:
                                numeric_value = float(value)
                                op[api_key] = locale.format_string("%.2f", numeric_value, grouping=True)
                            except (ValueError, TypeError):
                                op[api_key] = str(value) if value is not None else ''

                    all_records.append(op)

    if not all_records:
        return pd.DataFrame()

    df = pd.DataFrame(all_records)

    # Usa o MAP_COLUMNS do arquivo de configuração
    df = df.rename(columns=config.MAP_COLUMNS)

    # Usa o TARGET_COLUMNS do arquivo de configuração
    df_final = df.reindex(columns=config.TARGET_COLUMNS, fill_value='')

    return df_final


# --- Exportação para Excel ---

def export_to_excel(df, empresa_selecionada, empresa_code):
    """
    Exporta o DataFrame para um arquivo Excel com o layout, nome e formatação corretos.
    """
    if df.empty:
        messagebox.showinfo("Informação", "Nenhum dado para exportar.")
        return

    hoje = datetime.now()

    try:
        nome_empresa = empresa_selecionada.split(' - ')[1].upper()
    except:
        nome_empresa = "BASE"

    timestamp_formatado = hoje.strftime("%d_%m_%Y_%H_%M")
    initial_file = f"{nome_empresa}_ARQUIVO_BASE_800_{timestamp_formatado}.xlsx"

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialfile=initial_file
    )

    if not file_path: return

    try:
        dt_remessa_formatada = hoje.strftime("%d/%m/%Y")
        num_remessa = f"{empresa_code}{hoje.strftime('%y%m%d')}"

        # Definição de estilos
        none_border = Border()
        default_font = Font(bold=False)

        SHEET_NAME = 'Novas Operações'

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:

            if SHEET_NAME not in writer.book.sheetnames:
                writer.book.create_sheet(SHEET_NAME)
            sheet = writer.book[SHEET_NAME]

            # --- 1. Escreve os Títulos e Dados do DataFrame ---

            cols = list(df.columns)
            for c_idx, col_name in enumerate(cols, 1):
                cell = sheet.cell(row=3, column=c_idx, value=col_name)
                cell.border = none_border
                cell.font = default_font

            for r_idx, row in enumerate(dataframe_to_rows(df, header=False, index=False)):
                for c_idx, value in enumerate(row, 1):
                    cell = sheet.cell(row=r_idx + 4, column=c_idx, value=value)
                    cell.border = none_border

                    # --- 2. Escreve e formata o Cabeçalho de Controle (Linhas 1 e 2) ---

            sheet.cell(row=1, column=1, value="Dt. Remessa").border = none_border
            sheet.cell(row=1, column=2, value="Número da Remessa").border = none_border
            sheet.cell(row=1, column=3, value="Código da Empresa").border = none_border
            sheet.cell(row=1, column=4, value="Código de Evento Ref. A Atualização").border = none_border
            sheet.cell(row=1, column=5, value="Retomar/Liquidar Operacao não Presentes").border = none_border

            sheet.cell(row=2, column=1, value=dt_remessa_formatada).border = none_border
            sheet.cell(row=2, column=2, value=num_remessa).border = none_border
            sheet.cell(row=2, column=3, value=empresa_code).border = none_border
            sheet.cell(row=2, column=4, value="Ver: 07-05-2015").border = none_border
            sheet.cell(row=2, column=5, value="").border = none_border

            for r in range(1, 3):
                for c in range(1, 6):
                    sheet.cell(row=r, column=c).font = default_font

            sheet.column_dimensions['A'].width = 18
            sheet.column_dimensions['B'].width = 18
            sheet.column_dimensions['C'].width = 18

        messagebox.showinfo("Sucesso",
                            f"Dados exportados com sucesso para:\n{file_path}\n\nLembre-se de fazer a remoção manual das operações existentes via PROCV.")

    except Exception as e:
        messagebox.showerror("Erro de Exportação", f"Falha ao exportar para Excel: {e}")


# --- Interface Gráfica ---

class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Extrator de Operações Financeiras")
        self.geometry("480x420")
        self.resizable(False, False)
        self.configure(padx=10, pady=10)

        style = ttk.Style(self)
        style.configure('Gray.TLabel', foreground='gray')
        style.configure('TButton', font=('Arial', 10, 'bold'))

        self.create_widgets()

    def create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(padx=10, pady=10, fill='x')

        # --- Seleção da Empresa (usa EMPRESAS do config) ---
        ttk.Label(main_frame, text="1. Empresa (Banco):").grid(row=0, column=0, sticky='w', pady=5)
        self.empresa_var = tk.StringVar(self)
        self.empresa_var.set(list(config.EMPRESAS.keys())[0])
        self.empresa_menu = ttk.Combobox(main_frame, textvariable=self.empresa_var, values=list(config.EMPRESAS.keys()),
                                         state='readonly', width=30)
        self.empresa_menu.grid(row=0, column=1, columnspan=2, sticky='we', pady=5)

        # --- Parâmetros de Data (DateEntry - Calendário) ---

        venc_inicio_auto_str, venc_final_auto_str = calculate_dates()
        venc_inicio_auto = datetime.strptime(venc_inicio_auto_str, "%d/%m/%Y").date()
        venc_final_auto = datetime.strptime(venc_final_auto_str, "%d/%m/%Y").date()

        ttk.Label(main_frame, text="Vencimento Início:").grid(row=1, column=0, sticky='w', pady=2)
        self.venc_inicio_entry = DateEntry(main_frame, width=12, background='darkblue',
                                           foreground='white', borderwidth=2, date_pattern='dd/MM/yyyy')
        self.venc_inicio_entry.set_date(venc_inicio_auto)
        self.venc_inicio_entry.grid(row=1, column=1, sticky='w', pady=2, padx=2)

        ttk.Label(main_frame, text="Vencimento Final:").grid(row=2, column=0, sticky='w', pady=2)
        self.venc_final_entry = DateEntry(main_frame, width=12, background='darkblue',
                                          foreground='white', borderwidth=2, date_pattern='dd/MM/yyyy')
        self.venc_final_entry.set_date(venc_final_auto)
        self.venc_final_entry.grid(row=2, column=1, sticky='w', pady=2, padx=2)

        ttk.Label(main_frame, text="Use o ícone 📅 para selecionar a data.", style='Gray.TLabel').grid(row=3, column=0,
                                                                                                      columnspan=3,
                                                                                                      sticky='w',
                                                                                                      padx=5,
                                                                                                      pady=(0, 10))

        # --- Baixa Pdd ---
        ttk.Label(main_frame, text="2. Baixa Pdd (0/1):").grid(row=4, column=0, sticky='w', pady=5)
        self.baixa_pdd_var = tk.StringVar(self, value="0")
        self.baixa_pdd_menu = ttk.Combobox(main_frame, textvariable=self.baixa_pdd_var, values=["0", "1"],
                                           state='readonly', width=30)
        self.baixa_pdd_menu.grid(row=4, column=1, columnspan=2, sticky='we', pady=5)

        # --- Botão de Execução ---
        self.run_button = ttk.Button(self, text="3. Executar Processamento e Gerar Excel", command=self.execute_process)
        self.run_button.pack(pady=20, fill='x', padx=20)

        main_frame.grid_columnconfigure(1, weight=1)

    def update_status(self, text, is_error=False):
        """Atualiza o texto do botão e lida com o estado."""
        default_text = "3. Executar Processamento e Gerar Excel"
        self.run_button.config(text=text, state=tk.DISABLED if not is_error and text != default_text else tk.NORMAL)
        self.update()

    def execute_process(self):
        """Função principal que orquestra todo o processo."""
        self.update_status("Iniciando...")

        empresa_selecionada = self.empresa_var.get()
        empresa_key = config.EMPRESAS.get(empresa_selecionada)

        # 0. Checagem de Credenciais
        if not load_credentials():
            self.update_status("3. Executar Processamento e Gerar Excel")
            return

        try:
            banco_code = empresa_selecionada.split(' ')[0]
        except IndexError:
            self.update_status("Erro: Seleção de empresa inválida.", is_error=True)
            return

        venc_inicio = self.venc_inicio_entry.get()
        venc_final = self.venc_final_entry.get()
        baixa_pdd = self.baixa_pdd_var.get()

        # 1. Autenticação na API
        self.update_status("Autenticando na API...")
        token = authenticate_api(empresa_key)
        if not token:
            self.update_status("3. Executar Processamento e Gerar Excel")
            return

        # 2. Busca de dados na API
        self.update_status("Buscando dados na API...")
        api_data = fetch_api_data(empresa_key, venc_inicio, venc_final, baixa_pdd, token)
        if api_data is None:
            self.update_status("3. Executar Processamento e Gerar Excel")
            return

        if not api_data:
            messagebox.showinfo("Informação", "A API retornou uma lista de operações vazia.")
            self.update_status("3. Executar Processamento e Gerar Excel")
            return

        # 3. Processamento (aplica mapeamento e formatação)
        self.update_status("Processando e Formatando Dados...")
        df_final = process_data_no_db(api_data)

        if df_final.empty:
            messagebox.showinfo("Informação", "Nenhum dado válido para exportar foi encontrado na resposta da API.")
            self.update_status("3. Executar Processamento e Gerar Excel")
            return

        # 4. Exportação para Excel
        self.update_status(f"Exportando {len(df_final)} operações para Excel...")
        export_to_excel(df_final, empresa_selecionada, banco_code)

        # 5. Finalização
        self.update_status("3. Executar Processamento e Gerar Excel")


if __name__ == "__main__":
    try:
        app = Application()
        app.mainloop()
    except Exception as e:
        print(f"Erro fatal na aplicação: {e}", file=sys.stderr)
        messagebox.showerror("Erro Fatal", f"Ocorreu um erro inesperado e a aplicação será encerrada: {e}")